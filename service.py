import requests
from openpyxl import load_workbook


def read_sku_from_excel(path_to_file: str) -> list:
    """Парсинг xlsx файла"""
    sku_list = []

    wb = load_workbook(filename=path_to_file)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if any(row):
            sku_list.extend(row)

    return sku_list


def send_message_to_group(bot_token: str, chat_id: int, message: str):
    """Отправка уведомлений от бота в телеграм"""
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message}
    response = requests.post(url, json=payload)
    return response.json()
